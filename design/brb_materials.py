from typing import Any
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
from datetime import datetime


def generate_materials_excel(project_name, param_tables, project_folder=None, save_path=None):
 
    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BRB材料单"
    
    # 设置标题样式
    title_font = Font(name='宋体', size=16, bold=False)   # 标题字体
    content_font = Font(name='宋体', size=12, bold=False)  # 字体
    
    # 设置对齐方式
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)    # 居中对齐
    left_alignment = Alignment(horizontal='left', vertical='center') # 左对齐
    
    # 设置边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置白色背景样式
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # 1. 标题行
    title = f"{project_name}BRB材料单" # 项目名称后添加BRB材料单
    ws.merge_cells('A1:K1') # 合并A1到K1单元格，用于标题
    title_cell = ws['A1'] # 标题单元格
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
    # 给合并后的标题区域所有单元格加边框（仅左右上三边，去掉下边）和白色背景
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=11):
        for cell in row:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style=None)
            )
            cell.fill = white_fill
    ws.row_dimensions[1].height = 35 # 设置标题行高度

    # 设置宽度
    width_p = 0.62 # 列宽补偿值
    ws.column_dimensions['A'].width = 5.63+width_p
    ws.column_dimensions['B'].width = 8.38+width_p  
    ws.column_dimensions['C'].width = 7.25+width_p
    ws.column_dimensions['D'].width = 7.50+width_p
    ws.column_dimensions['E'].width = 4.5+width_p
    ws.column_dimensions['F'].width = 5.5+width_p
    ws.column_dimensions['G'].width = 8.38+width_p
    ws.column_dimensions['H'].width = 5+width_p
    ws.column_dimensions['I'].width = 8.75+width_p
    ws.column_dimensions['J'].width = 11+width_p
    ws.column_dimensions['K'].width = 8.5+width_p

    # 第二行增加时间
    ws.merge_cells('A2:K2') # 合并A2到K2单元格，用于时间
    time_cell = ws['A2'] # 时间单元格
    time_cell.value = f"{datetime.now().strftime('%Y年%m月%d日')}"
    time_cell.font = content_font
    time_cell.alignment = Alignment(horizontal='right', vertical='center')
    # 给合并后的时间区域所有单元格加边框（仅左右下三边，去掉上边）和白色背景
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=11):
        for cell in row:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style=None),
                bottom=Side(style='thin')
            )
            cell.fill = white_fill
    ws.row_dimensions[2].height = 15 # 设置时间行高度
    
    # 2. 材料清单表格 

    # 设置起点
    start_row = 3 # 初始值为3，从第3行开始填充数据
    # 统计材料规格，初始化
        # 板厚列表和材料列表
    think_list = [("Q235",6)] # 板厚列表，初始值为6mm,（方板厚度），默认材料为Q235
        # 方管宽度和厚度列表
    pipe_width_think_list = []

    
    # 填写参数表
    for table in param_tables: # 遍历每个参数表
        think = int(table.get('params', {}).get('板材厚度(mm)', ''))
        core_material = table.get('core_material', 'Q235')
        print(core_material)
        width = int(table.get('params', {}).get('截面宽度(mm)', ''))
        height = int(table.get('params', {}).get('截面高度(mm)', ''))
        # weld_height = int(table.get('params', {}).get('焊缝高度(mm)', ''))
        pipe_width = int(table.get('params', {}).get('方管宽度(mm)', ''))
        pipe_thickness = int(table.get('params', {}).get('方管厚度(mm)', '4'))  # 增加方管厚度，默认值4mm
        length_quantity = table.get('length_quantity', [])    
        design_force = table.get('design_force', '')

        think_list.append((core_material,think)) # 板厚列表添加当前板厚和材料
        think_list = list[tuple[str,int]](set[tuple[str,int]](think_list)) # 去除重复元素

        pipe_width_think_list.append((pipe_width,pipe_thickness)) # 方管宽度和厚度列表添加当前方管宽度和厚度
        pipe_width_think_list = list[tuple[int,int]](set[tuple[int,int]](pipe_width_think_list)) # 去除重复元素

        for lq in length_quantity:
            length = lq[0] if isinstance(lq, tuple) else lq.get('length', '')
            quantity = lq[1] if isinstance(lq, tuple) else lq.get('quantity', '')             

            # 2.1 表头
            headers1 = ["序号", "型号", "名称", "规格", "数量", "单重(Kg)", "总重(Kg)", "备注"]
            headers2 = ["屈服荷载(KN)",  "长度(mm)","厚","宽","长",]

            # 设置第一行表头
            ws.cell(row=start_row, column=1, value=headers1[0])
            ws.cell(row=start_row, column=2, value=headers1[1])
            ws.cell(row=start_row, column=4, value=headers1[2])
            ws.cell(row=start_row, column=5, value=headers1[3])
            ws.cell(row=start_row, column=8, value=headers1[4])
            ws.cell(row=start_row, column=9, value=headers1[5])
            ws.cell(row=start_row, column=10, value=headers1[6])
            ws.cell(row=start_row, column=11, value=headers1[7])       
            ws.row_dimensions[start_row].height = 15 # 第一行表头高度

            # 设置第二行表头
            ws.cell(row=start_row+1, column=2, value=headers2[0])
            ws.cell(row=start_row+1, column=3, value=headers2[1])
            ws.cell(row=start_row+1, column=5, value=headers2[2])
            ws.cell(row=start_row+1, column=6, value=headers2[3])
            ws.cell(row=start_row+1, column=7, value=headers2[4])
            ws.row_dimensions[start_row+1].height = 28.5 # 第二行表头高度

            # 合并表头单元格
            ws.merge_cells(f'A{start_row}:A{start_row+1}') # 合并A2到A3单元格
            ws.merge_cells(f'B{start_row}:C{start_row}') # 合并B2到C2单元格
            ws.merge_cells(f'D{start_row}:D{start_row+1}') # 合并D2到D3单元格
            ws.merge_cells(f'E{start_row}:G{start_row}') # 合并E2到G2单元格
            ws.merge_cells(f'H{start_row}:H{start_row+1}') # 合并H2到H3单元格
            ws.merge_cells(f'I{start_row}:I{start_row+1}') # 合并I2到I3单元格
            ws.merge_cells(f'J{start_row}:J{start_row+1}') # 合并J2到J3单元格
            ws.merge_cells(f'K{start_row}:K{start_row+1}') # 合并K2到K3单元格
          
            # 设置表头字体和对齐方式
            for row in range(start_row, start_row+2):
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.font = content_font
                    cell.alignment = center_alignment

            # 2.2 内容
            # 读取截面类型
            template_type = table.get('template_type')
            print(template_type)

            start_row += 2 # 表格起始行号跨表头两行


            # 根据截面类型，填充相关参数，并更改插入位置
            if template_type == '十一':
                num_rows = 8 # 表格总行数

                # 设置单元格样式
                for row in range(start_row, start_row+num_rows):
                    ws.row_dimensions[row].height = 14.25 # 设置行高度为14.25
                    for col in range(1, 12):
                        cell = ws.cell(row=row, column=col)
                        cell.font = content_font  # 设置内容字体为10号字体
                        cell.alignment = center_alignment # 设置内容居中对齐
                    
                
                # 填充A列序号
                for row in range(start_row, start_row+4):
                    ws.cell(row=row, column=1, value=f"{row-start_row+1}")

                ws.cell(row=start_row+4, column=1, value="钢材合计：")
                ws.merge_cells(f'A{start_row+4}:G{start_row+4}') # 合并单元格

                for row in range(start_row+5, start_row+7):
                    ws.cell(row=row, column=1, value=f"{row-start_row}")

                # 填充型号
                ws.cell(row=start_row, column=2, value=design_force) # 设计力
                ws.cell(row=start_row, column=3, value=length) # 长度
                ws.cell(row=start_row+5, column=2, value="灌浆")
                ws.cell(row=start_row+6, column=2, value="合计")

                # 填充名称
                ws.cell(row=start_row, column=4, value="芯板")
                ws.cell(row=start_row+1, column=4, value="翼缘板")
                ws.cell(row=start_row+2, column=4, value="挡板")
                ws.cell(row=start_row+3, column=4, value="方管")

                # 填充厚度
                ws.cell(row=start_row, column=5, value=think) # 芯板厚度
                ws.cell(row=start_row+1, column=5, value=think) # 翼缘板1厚度
                ws.cell(row=start_row+2, column=5, value=6) # 挡板厚度
                ws.cell(row=start_row+3, column=5, value=pipe_thickness) # 方管厚度
                ws.cell(row=start_row+5, column=5, value=pipe_width) # 灌浆厚度

                # 填充宽度
                ws.cell(row=start_row, column=6, value=height) # 芯板宽度
                ws.cell(row=start_row+1, column=6, value=(width- think)/2) # 翼缘板宽度
                ws.cell(row=start_row+2, column=6, value=pipe_width+20) # 挡板宽度
                ws.cell(row=start_row+3, column=6, value=pipe_width*4) # 方管宽度
                ws.cell(row=start_row+5, column=6, value=pipe_width) # 灌浆宽度

                # 填充长度
                ws.cell(row=start_row, column=7, value=f"=C{start_row}") # 芯板长度
                ws.cell(row=start_row+1, column=7, value=f"300") # 翼缘板长度
                ws.cell(row=start_row+2, column=7, value=f"=f{start_row+3}") # 挡板长度
                ws.cell(row=start_row+3, column=7, value=f"=C{start_row}-300") # 方管长度
                ws.cell(row=start_row+5, column=7, value=f"=C{start_row}-300") # 灌浆长度
                
                #填充数量
                ws.cell(row=start_row, column=8, value=quantity) #芯板数量
                ws.cell(row=start_row+1, column=8, value=quantity*4) #翼缘板数量
                ws.cell(row=start_row+2, column=8, value=quantity*2) #挡板数量
                ws.cell(row=start_row+3, column=8, value=quantity*1) #方管数量
                ws.cell(row=start_row+5, column=8, value=quantity*1) # 灌浆数量

                # 填充单重,保留两位小数
                for row in range(start_row, start_row+4):
                    ws.cell(row=row, column=9, value=f"=E{row}*F{row}*G{row}*{0.00000785:.8f}").number_format = '0.00' # 芯板数量
                ws.cell(row=start_row+5, column=9, value=f"=E{start_row+6}*F{start_row+6}*G{start_row+6}*{0.0000021:.8f}").number_format = '0.00' # 灌浆数量

                # 填充总重
                for row in range(start_row, start_row+4):
                    ws.cell(row=row, column=10, value=f"=I{row}*H{row}").number_format = '0.00' # 芯板总数量
                ws.cell(row=start_row+5, column=10, value=f"=I{start_row+6}*H{start_row+6}").number_format = '0.00' # 灌浆总数量 

                # 合计重量
                ws.cell(row=start_row+4, column=10, value=f"=SUM(J{start_row}:J{start_row+3})").number_format = '0.00' # 钢材总重
                ws.cell(row=start_row+6, column=10, value=f"=SUM(J{start_row+4}:J{start_row+5})").number_format = '0.00' # 钢材加灌浆总重量
                ws.cell(row=start_row+4, column=9, value=f"=J{start_row+4}/{quantity}").number_format = '0.00' # 单个BRB重量

                # 备注
                for row in range(start_row, start_row+3):
                    ws.cell(row=row, column=11, value=f"{core_material}" if ws.cell(row=row, column=4 ).value != "挡板" else "Q235") # 材料
                ws.merge_cells(f'K{start_row+4}:K{start_row+6}') # 合并单元格
                ws.cell(row=start_row+4, column=11, value=f"十字") # 截面
                
                # 所有单元格加边框
                for row in ws.iter_rows(min_row=start_row-2, max_row=start_row+6, min_col=1, max_col=11):
                    for cell in row:
                        cell.border = thin_border

                # 移动插入点到末尾
                start_row += num_rows

            else:# 王工和王一
                num_rows = 9 # 参数总行数
                # 设置单元格样式
                for row in range(start_row, start_row+num_rows):
                    ws.row_dimensions[row].height = 14.25 # 设置行高度为14.25
                    for col in range(1, 12):
                        cell = ws.cell(row=row, column=col)
                        cell.font = content_font  # 设置内容字体为10号字体
                        cell.alignment = center_alignment # 设置内容居中对齐                   
                
                # 填充A列序号
                for row in range(start_row, start_row+5):
                    ws.cell(row=row, column=1, value=f"{row-start_row+1}")
                
                ws.cell(row=start_row+5, column=1, value="钢材合计：")
                ws.merge_cells(f'A{start_row+5}:G{start_row+5}') # 合并单元格

                for row in range(start_row+6, start_row+8):
                    ws.cell(row=row, column=1, value=f"{row-start_row}")

                # 填充型号
                ws.cell(row=start_row, column=2, value=design_force) # 设计力
                ws.cell(row=start_row, column=3, value=length) # 长度
                ws.cell(row=start_row+6, column=2, value="灌浆")
                ws.cell(row=start_row+7, column=2, value="合计")

                # 填充名称
                ws.cell(row=start_row, column=4, value="芯板")
                ws.cell(row=start_row+1, column=4, value="翼缘板1")
                ws.cell(row=start_row+2, column=4, value="翼缘板2")
                ws.cell(row=start_row+3, column=4, value="挡板")
                ws.cell(row=start_row+4, column=4, value="方管")

                # 填充厚度
                ws.cell(row=start_row, column=5, value=think) # 芯板厚度
                ws.cell(row=start_row+1, column=5, value=think) # 翼缘板1厚度
                ws.cell(row=start_row+2, column=5, value=think) # 翼缘板2厚度
                ws.cell(row=start_row+3, column=5, value=6) # 挡板厚度
                ws.cell(row=start_row+4, column=5, value=pipe_thickness) # 方管厚度
                ws.cell(row=start_row+6, column=5, value=pipe_width) # 灌浆厚度

                # 填充宽度
                ws.cell(row=start_row, column=6, value=height- think- think) # 芯板宽度
                ws.cell(row=start_row+1, column=6, value=width) # 翼缘板1宽度
                ws.cell(row=start_row+2, column=6, value=(width- think)/2) # 翼缘板2宽度
                ws.cell(row=start_row+3, column=6, value=pipe_width+20) # 挡板宽度
                ws.cell(row=start_row+4, column=6, value=pipe_width*4) # 方管宽度
                ws.cell(row=start_row+6, column=6, value=pipe_width) # 灌浆宽度

                # 填充长度
                ws.cell(row=start_row, column=7, value=f"=C{start_row}") # 芯板长度
                ws.cell(row=start_row+1, column=7, value=f"300" if template_type =='王一' else f"=C{start_row}") # 翼缘板1长度
                ws.cell(row=start_row+2, column=7, value=f"300") # 翼缘板2长度
                ws.cell(row=start_row+3, column=7, value=f"=f{start_row+3}") # 挡板长度
                ws.cell(row=start_row+4, column=7, value=f"=C{start_row}-300") # 方管长度
                ws.cell(row=start_row+6, column=7, value=f"=C{start_row}-300") # 灌浆长度
                
                #填充数量
                ws.cell(row=start_row, column=8, value=quantity) #芯板数量
                ws.cell(row=start_row+1, column=8, value=quantity*4 if template_type =='王一' else quantity*2) #翼缘板1数量
                ws.cell(row=start_row+2, column=8, value=(quantity*4)) #翼缘板2数量
                ws.cell(row=start_row+3, column=8, value=quantity*2) #挡板数量
                ws.cell(row=start_row+4, column=8, value=quantity*1) #方管数量
                ws.cell(row=start_row+6, column=8, value=quantity*1) # 灌浆数量

                # 填充单重,保留两位小数
                for row in range(start_row, start_row+5):
                    ws.cell(row=row, column=9, value=f"=E{row}*F{row}*G{row}*{0.00000785:.8f}").number_format = '0.00' # 芯板数量
                ws.cell(row=start_row+6, column=9, value=f"=E{start_row+6}*F{start_row+6}*G{start_row+6}*{0.0000021:.8f}").number_format = '0.00' # 灌浆数量

                # 填充总重
                for row in range(start_row, start_row+5):
                    ws.cell(row=row, column=10, value=f"=I{row}*H{row}").number_format = '0.00' # 芯板总数量
                ws.cell(row=start_row+6, column=10, value=f"=I{start_row+6}*H{start_row+6}").number_format = '0.00' # 灌浆总数量 

                # 合计重量
                ws.cell(row=start_row+5, column=10, value=f"=SUM(J{start_row}:J{start_row+4})").number_format = '0.00' # 钢材总重
                ws.cell(row=start_row+7, column=10, value=f"=SUM(J{start_row+5}:J{start_row+6})").number_format = '0.00' # 钢材加灌浆总重量
                ws.cell(row=start_row+5, column=9, value=f"=J{start_row+5}/{quantity}").number_format = '0.00' # 单个BRB重量

                # 备注
                for row in range(start_row, start_row+4):
                    ws.cell(row=row, column=11, value=f"{core_material}" if ws.cell(row=row, column=4 ).value != "挡板" else "Q235") # 材料
                ws.merge_cells(f'K{start_row+5}:K{start_row+7}') # 合并单元格
                ws.cell(row=start_row+5, column=11, value=f"王字") # 截面

                 # 所有单元格加边框
                for row in ws.iter_rows(min_row=start_row-2, max_row=start_row+7, min_col=1, max_col=11):
                    for cell in row:
                        cell.border = thin_border

                # 移动插入点到末尾
                start_row += num_rows
            
    # 3. 制作汇总表格
    start_row += 1 # 移动插入点到下一行
    # 3.1 汇总表格标题
    title = f"{project_name}BRB材料汇总" # 项目名称后添加BRB材料汇总
    ws.merge_cells(f'A{start_row}:I{start_row}') # 合并A1到I1单元格，用于标题
    title_cell = ws[f'A{start_row}'] # 标题单元格
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    # 给合并后的标题区域所有单元格加边框
    for row in ws.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = thin_border
    ws.row_dimensions[start_row].height = 25 # 设置标题行高度
    start_row += 1 # 移动插入点到下一行

    # 3.2 汇总表格表头
    headers3 = ["序号", "材料型号", "总重(kg)", "备注"]
    ws.cell(row=start_row, column=1, value=headers3[0])
    ws.cell(row=start_row, column=2, value=headers3[1])
    ws.cell(row=start_row, column=4, value=headers3[2])
    ws.cell(row=start_row, column=6, value=headers3[3])
        # 合并表头单元格
    ws.merge_cells(f'B{start_row}:C{start_row}') # 合并B到C单元格
    ws.merge_cells(f'D{start_row}:E{start_row}') # 合并D到E单元格
    ws.merge_cells(f'F{start_row}:I{start_row}') # 合并F到I单元格

    start_row += 1 # 移动插入点到下一行

    # 3.3 汇总表格数据填写
    think_list.sort() # 板厚列表排序
    pipe_width_think_list.sort() # 方管宽度和厚度列表排序
    print(think_list) # 打印板厚列表
    print(pipe_width_think_list) # 打印方管宽度和厚度列表

    # 3.3.1 填写板厚数据
    for idx, (core_material,think) in enumerate(think_list, start=1):
        ws.cell(row=start_row, column=1, value=idx)  # 序号
        ws.cell(row=start_row, column=2, value=think).number_format = '0"mm钢板"'
        ws.cell(row=start_row, column=4, value=f"=SUMIFS($J:$J,$E:$E,B{start_row},$K:$K,F{start_row})").number_format = '0.00' # 总重
        ws.cell(row=start_row, column=6, value=f"{core_material}") # 备注
        # 合并单元格
        ws.merge_cells(f'B{start_row}:C{start_row}') # 合并B到C单元格
        ws.merge_cells(f'D{start_row}:E{start_row}') # 合并D到E单元格
        ws.merge_cells(f'F{start_row}:I{start_row}') # 合并F到I单元格

        start_row += 1 # 移动插入点到下一行

    # 3.3.2 填写方管数据
    for idx, (pipe_width, think) in enumerate(pipe_width_think_list, start=idx+1):
        ws.cell(row=start_row, column=1, value=idx)  # 序号
        ws.cell(row=start_row, column=2, value=pipe_width).number_format = '0"方管"' # 材料型号
        ws.cell(row=start_row, column=3, value=think).number_format = '0"mm厚"' # 材料型号
        ws.cell(row=start_row, column=4, value=f'=SUMIFS($J:$J,$E:$E,C{start_row},$F:$F,B{start_row}*4,$D:$D,"方管")').number_format = '0.00' # 总重
        ws.cell(row=start_row, column=6, value=f"=D{start_row}/(B{start_row}*4*C{start_row})/0.00000785").number_format = '"总长"0"mm"' # 备注
        ws.cell(row=start_row, column=8, 
            value=f"=IF(ROUNDUP(F{start_row}/12000,0)<3,\
            ROUNDUP(F{start_row}/12000,0),\
            IF((ROUNDUP(F{start_row}/12000,0)*12000-F{start_row})<300,(ROUNDUP(F{start_row}/12000,0)+1),ROUNDUP(F{start_row}/12000,0)))")\
                .number_format = '"备"0"根12米"' # 备注
        # 合并单元格
        ws.merge_cells(f'D{start_row}:E{start_row}') # 合并D到E单元格
        ws.merge_cells(f'F{start_row}:G{start_row}') # 合并F到G单元格
        ws.merge_cells(f'H{start_row}:I{start_row}') # 合并H到I单元格
        
        start_row += 1 # 移动插入点到下一行
    
    # 3.3.3 钢材合计
    ws.cell(row=start_row, column=1, value="钢材合计")  
    ws.cell(row=start_row, column=4, value=f"=SUM(D{start_row-idx}:D{start_row-1})").number_format = '0.00' # 总重
        # 合并单元格
    ws.merge_cells(f'A{start_row}:C{start_row}') # 合并A到C单元格
    ws.merge_cells(f'D{start_row}:E{start_row}') # 合并D到E单元格
    ws.merge_cells(f'F{start_row}:I{start_row}') # 合并F到I单元格
    start_row += 1 # 移动插入点到下一行

    # 3.3.4 灌浆
    idx += 1
    ws.cell(row=start_row, column=1, value=idx)  
    ws.cell(row=start_row, column=2, value=f"灌浆") # 材料型号
    ws.cell(row=start_row, column=4, value=f"=SUMIF($B:$B,B{start_row},$J:$J)").number_format = '0.00' # 总重
        # 合并单元格
    ws.merge_cells(f'B{start_row}:C{start_row}') # 合并B到C单元格
    ws.merge_cells(f'D{start_row}:E{start_row}') # 合并D到E单元格
    ws.merge_cells(f'F{start_row}:I{start_row}') # 合并F到I单元格
    start_row += 1 # 移动插入点到下一行

    # 3.3.5 合计
    ws.cell(row=start_row, column=1, value="合计")  
    ws.cell(row=start_row, column=4, value=f"=SUM(D{start_row-2}:D{start_row-1})").number_format = '0.00' # 总重
    ws.cell(row=start_row, column=6, value=f'=SUMIF($D:$D,"芯板",$H:$H)').number_format = '"共"0"件BRB产品"' # 总重
        # 合并单元格
    ws.merge_cells(f'B{start_row}:C{start_row}') # 合并B到C单元格
    ws.merge_cells(f'D{start_row}:E{start_row}') # 合并D到E单元格
    ws.merge_cells(f'F{start_row}:I{start_row}') # 合并F到I单元格

    # 3.3.6 设置格式
    # 设置字体和对齐方式
    for row in range(start_row-idx-2, start_row+1):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.font = content_font
            cell.alignment = center_alignment
    # 设置边框
    for row in range(start_row-idx-2, start_row+1):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border


    start_row += 1
    
    # 根据参数决定是否保存到磁盘
    if save_path:
        # 如果提供了保存路径，直接保存到该路径
        wb.save(save_path)
        return save_path
    else:
        # 否则，将文件保存到内存中
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# 测试代码
if __name__ == "__main__":
    # 测试数据
    test_project_name = "测试项目"
    test_param_tables = [
        {
            "table_number": 1,
            "template_type": "王一",
            "design_force": "2300",
            "core_material": "Q235",
            "params": 
            {
                "截面宽度(mm)": 170,
                "截面高度(mm)": 170,
                "板材厚度(mm)": 16,
                "焊缝高度(mm)": 12,
                "方管宽度(mm)": 250,
                "方管厚度(mm)": 5
            },
            "length_quantity": [
                {"length": 7800, "quantity": 2},
                {"length": 1500, "quantity": 3}
            ]
        },
        {
            "table_number": 2,
            "design_force": "3000",
            "core_material": "Y160",
            "template_type": "十一",
            "params": {
                "截面宽度(mm)": 180,
                "截面高度(mm)": 180,
                "板材厚度(mm)": 12,
                "焊缝高度(mm)": 10,
                "方管宽度(mm)": 140,
                "方管厚度(mm)": 4
            },
            "length_quantity": [
                {"length": 2000, "quantity": 4}
            ]
        },
        {
            "table_number": 3,
            "design_force": "4000",
            "core_material": "Q235",
            "template_type": "王工",
            "params": {
                "截面宽度(mm)": 200,
                "截面高度(mm)": 200,
                "板材厚度(mm)": 14,
                "焊缝高度(mm)": 12,
                "方管宽度(mm)": 160,
                "方管厚度(mm)": 6
            },
            "length_quantity": [
                {"length": 2500, "quantity": 2},
                {"length": 3000, "quantity": 1}

            ]
        }
    ]
    
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 生成测试Excel文件，保存到脚本所在目录
    save_file_path = os.path.join(script_dir, f"{test_project_name}_BRB材料单.xlsx")
    excel_path = generate_materials_excel(test_project_name, test_param_tables, save_path=save_file_path)
    print(f"测试材料单已生成：{excel_path}")
